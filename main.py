import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment
from aiogram import Bot, Dispatcher, executor, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup
from aiogram.types import InlineKeyboardMarkup, InlineKeyboardButton, ContentType
from aiogram.utils import executor
from aiogram.dispatcher.filters import Text
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton
from typing import List

API_TOKEN = '6359219248:AAH9bqLUP0MnvSYpnsKjcJUs7PQ1uNHmoWM'  # Replace with your token

ADMIN_IDS: List[int] = []  # Populate with admin user IDs

logging.basicConfig(level=logging.INFO)

welcome_start = False

# States
# Define states
bot = Bot(token=API_TOKEN)
storage = MemoryStorage()
dp = Dispatcher(bot, storage=storage)

class MenuState(StatesGroup):
    product_info = State()

class QuestionState(StatesGroup):
    waiting_for_contact = State()
    waiting_for_question_text = State()
    waiting_for_media = State()

class ReportProblemState(StatesGroup):
    waiting_for_contact = State()
    waiting_for_marketplace_choice = State()
    waiting_for_message = State()
    waiting_for_media = State()


# Read admin IDs from a text file
def load_admin_ids():
    try:
        with open("admin_ids.txt", "r") as f:
            for line in f:
                ADMIN_IDS.append(int(line.strip()))
    except FileNotFoundError:
        logging.warning("admin_ids.txt file not found, no admins loaded.")

load_admin_ids()

@dp.message_handler(lambda message: message.text == "🌚🌚🌚")
async def add_new_admin(message: types.Message):
    new_admin_id = message.from_user.id
    with open("admin_ids.txt", "a+") as f:
        f.seek(0)  # Переместить указатель в начало файла для чтения
        admin_ids = f.read().splitlines()
        if str(new_admin_id) not in admin_ids:  # Проверить, что админ не добавлен ранее
            f.write(f"{new_admin_id}\n")  # Добавить новый ID в файл
            await message.reply("Новый админ успешно добавлен.")
            ADMIN_IDS.append(new_admin_id)  # Опционально: добавить в список админов в памяти
        else:
            await message.reply("Этот пользователь уже является администратором.")


@dp.message_handler(commands=['excel'])
async def send_excel_files(message: types.Message):
    # Отправка файла problems.xlsx
    with open('problems.xlsx', 'rb') as problems_file:
        await message.answer_document(problems_file)

    # Отправка файла questions.xlsx
    with open('questions.xlsx', 'rb') as questions_file:
        await message.answer_document(questions_file)


@dp.message_handler(commands=['start'], state='*')
async def send_welcome_and_show_main_menu(message: types.Message, state: FSMContext):
    global welcome_start
    # Сбрасываем состояние для избежания конфликтов
    await state.reset_state(with_data=True)

    if welcome_start == False:
        # Сообщение приветствия
        welcome_message = (
            "Добро пожаловать в бот службы поддержки бренда IHI. Мы очень дорожим качеством нашей продукции и "
            "отношениями с клиентами. Мы всегда готовы помочь, если возникли какие-либо трудности. С помощью кнопок "
            "ниже вы сможете ознакомиться с линейкой продуктов IHI, задать вопрос, а также сообщить о возникшей проблеме."
        )
        await message.answer(welcome_message)

    welcome_start = True

    # Показываем главное меню
    keyboard = InlineKeyboardMarkup(row_width=1)
    buttons = [
        InlineKeyboardButton("📦 Продукция IHI", callback_data='product_info'),
        InlineKeyboardButton("🤔 Задать вопрос", callback_data='ask_question'),
        InlineKeyboardButton("⚠️ Сообщить о проблеме", callback_data='report_problem')
    ]
    keyboard.add(*buttons)
    await message.answer("Выберите действие:", reply_markup=keyboard)


@dp.callback_query_handler(lambda c: c.data == 'product_info', state='*')
async def handle_main_menu(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)

    await MenuState.product_info.set()  # Переходим в состояние product_info

    keyboard = InlineKeyboardMarkup(row_width=1)
    buttons = [
        InlineKeyboardButton("🛒 OZON", callback_data='ozon'),
        InlineKeyboardButton("🛍️ Wildberries", callback_data='wildberries'),
        InlineKeyboardButton("🔙 Назад", callback_data='back')
    ]
    keyboard.add(*buttons)
    await bot.send_message(callback_query.from_user.id, "Выберите площадку:", reply_markup=keyboard)

# Исправленный обработчик для действий внутри состояния product_info
# Исправленный обработчик для действий внутри состояния product_info
@dp.callback_query_handler(state=MenuState.product_info)
async def handle_product_info(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    action = callback_query.data

    if action == 'ozon':
        await bot.send_message(callback_query.from_user.id, "Пожалуйста, перейдите по ссылке для просмотра продукции IHI на OZON:\n\nhttps://www.ozon.ru/brand/ihi-100869869/")
        await send_welcome_and_show_main_menu(callback_query.message, state)
    elif action == 'wildberries':
        await bot.send_message(callback_query.from_user.id, "Пожалуйста, перейдите по ссылке для просмотра продукции IHI на WildBerries:\n\nhttps://www.wildberries.ru/brands/910076-IHI")
        await send_welcome_and_show_main_menu(callback_query.message, state)
    elif action == "back":
        await send_welcome_and_show_main_menu(callback_query.message, state)

    await state.finish()


# Handler for "Ask a question" callback - assumes implementation is within the state context
@dp.callback_query_handler(lambda c: c.data == 'ask_question')
async def ask_for_contact(callback_query: types.CallbackQuery, state: FSMContext):
    await QuestionState.waiting_for_contact.set()
    # Создание клавиатуры для запроса контакта
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    contact_button = KeyboardButton("Отправить мой контакт 📞", request_contact=True)
    keyboard.add(contact_button)
    await bot.send_message(callback_query.from_user.id,
                           "Пожалуйста, используйте кнопку ниже, чтобы отправить ваш контакт.",
                           reply_markup=keyboard)


# Handling the back button functionality correctly to return to the main menu from any state
@dp.callback_query_handler(text='back_to_menu', state='*')
async def back_to_main_menu(callback_query: types.CallbackQuery, state: FSMContext):
    await state.reset_state()

# Adjusting the contact received handler to capture contact and ask for the question text
@dp.message_handler(content_types=types.ContentType.CONTACT, state=QuestionState.waiting_for_contact)
async def contact_received(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.contact.phone_number)
    await QuestionState.waiting_for_question_text.set()
    await message.reply("Пожалуйста, введите ваш вопрос.")

# Processing the question text and asking for optional media
@dp.message_handler(state=QuestionState.waiting_for_question_text)
async def question_text_received(message: types.Message, state: FSMContext):
    await state.update_data(message=message.text)
    skip_button = InlineKeyboardButton("Пропустить шаг", callback_data='skip_media')
    keyboard = InlineKeyboardMarkup().add(skip_button)
    await QuestionState.waiting_for_media.set()
    await message.reply("Вы можете приложить фото или видео. Если не хотите прикладывать медиа, нажмите 'Пропустить шаг'.", reply_markup=keyboard)

@dp.callback_query_handler(text='skip_media', state=QuestionState.waiting_for_media)
async def skip_media(callback_query: types.CallbackQuery, state: FSMContext):
    await bot.answer_callback_query(callback_query.id)
    await send_data_to_admins(callback_query.message, state)
    await state.finish()

@dp.message_handler(content_types=['photo', 'video'], state=QuestionState.waiting_for_media)
async def media_received(message: types.Message, state: FSMContext):
    await send_data_to_admins(message, state, is_media=True)
    await state.finish()

async def send_data_to_admins(message: types.Message, state: FSMContext, is_media: bool = False):
    user_data = await state.get_data()
    print("тут")
    # Проверяем, есть ли у пользователя username
    username = message.from_user.username
    user_display_name = f"@{username}" if username else f"{message.from_user.first_name} {message.from_user.last_name}".strip()

    for admin_id in ADMIN_IDS:
        await bot.send_message(admin_id,
                               f"Пользователь {user_display_name} отправил следующие данные из раздела '🤔 Задать вопрос':\n"
                               f"Номер телефона: {user_data['phone']}\n"
                               f"Сообщение: {user_data.get('message', 'No message provided')}\n")
        if is_media:
            if message.photo:
                await bot.send_photo(admin_id, photo=message.photo[-1].file_id)
            elif message.video:
                await bot.send_video(admin_id, video=message.video[-1].file_id)

    # Сообщение пользователю о успешной отправке
    welcome_message = (
        "Ваш вопрос успешно отправлен. Благодарим за обратную связь. Наш менеджер свяжется с Вами в ближайшее время! С Уважением, Команда поддержки бренда IHI!"
    )
    await message.answer(welcome_message)
    await send_welcome_and_show_main_menu(message, state)

    # Сохранение данных в Excel файл
    try:
        wb = openpyxl.load_workbook('questions.xlsx')  # Попытка загрузить существующий файл
    except FileNotFoundError:
        wb = Workbook()  # Если файл не найден, создаем новый
        ws = wb.active
        ws.append(['Пользователь', 'Номер телефона', 'Сообщение'])
    else:
        ws = wb.active

    row_data = [user_display_name, user_data['phone'], user_data.get('message', 'No message provided')]
    ws.append(row_data)

    wb.save('questions.xlsx')






@dp.callback_query_handler(lambda c: c.data == 'report_problem')
async def report_problem_handler(callback_query: types.CallbackQuery, state: FSMContext):
    # Переходим к состоянию ожидания контакта без предварительных условий
    await ReportProblemState.waiting_for_contact.set()
    contact_button = KeyboardButton("Отправить мой контакт 📞", request_contact=True)
    keyboard = ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True).add(contact_button)
    await bot.send_message(callback_query.from_user.id,
                           "Пожалуйста, используйте кнопку ниже, чтобы отправить ваш контакт.",
                           reply_markup=keyboard)
    await bot.answer_callback_query(callback_query.id)

@dp.message_handler(content_types=ContentType.CONTACT, state=ReportProblemState.waiting_for_contact)
async def contact_received(message: types.Message, state: FSMContext):
    await state.update_data(phone=message.contact.phone_number)

    buttons = [
        InlineKeyboardButton("Ozon", callback_data='ozon'),
        InlineKeyboardButton("Wildberries", callback_data='wildberries')
    ]
    keyboard = InlineKeyboardMarkup(inline_keyboard=[buttons])
    await message.answer("Выберите маркетплейс где покупали товар:", reply_markup=keyboard)
    await ReportProblemState.waiting_for_marketplace_choice.set()

@dp.callback_query_handler(state=ReportProblemState.waiting_for_marketplace_choice)
async def marketplace_choice_received(callback_query: types.CallbackQuery, state: FSMContext):
    await state.update_data(marketplace=callback_query.data)
    await callback_query.message.answer("Введите ваше сообщение.")
    await ReportProblemState.waiting_for_message.set()

@dp.callback_query_handler(text='skip_step', state=ReportProblemState.waiting_for_message)
async def skip_media_step(callback_query: types.CallbackQuery, state: FSMContext):
    await send_data_to_admins_two(callback_query.message, state)
    await state.finish()

@dp.message_handler(state=ReportProblemState.waiting_for_message, content_types=ContentType.TEXT)
async def text_message_received(message: types.Message, state: FSMContext):
    await state.update_data(message=message.text)
    skip_button = InlineKeyboardButton("Пропустить", callback_data='skip_media_upload')
    keyboard = InlineKeyboardMarkup().add(skip_button)
    await message.answer("Прикрепите фото или видео, если необходимо. Или пропустите этот шаг.", reply_markup=keyboard)
    await ReportProblemState.waiting_for_media.set()

@dp.callback_query_handler(text='skip_media_upload', state=ReportProblemState.waiting_for_media)
async def skip_media_upload(callback_query: types.CallbackQuery, state: FSMContext):
    await send_data_to_admins_two(callback_query.message, state)
    await state.finish()

@dp.message_handler(content_types=[ContentType.PHOTO, ContentType.VIDEO], state=ReportProblemState.waiting_for_media)
async def media_received(message: types.Message, state: FSMContext):
    if message.content_type == ContentType.PHOTO:
        media_id = message.photo[-1].file_id
    else:  # ContentType.VIDEO
        media_id = message.video.file_id
    await state.update_data(media_id=media_id)
    await send_data_to_admins_two(message, state, is_media=True)
    await state.finish()

async def send_data_to_admins_two(message: types.Message, state: FSMContext, is_media: bool = False):
    user_data = await state.get_data()
    marketplace = user_data.get('marketplace', 'Не указан')
    phone = user_data.get('phone', 'Не указан')
    user_message = user_data.get('message', 'Сообщение не предоставлено')
    media_id = user_data.get('media_id', None)

    # Проверяем, есть ли у пользователя username
    username = message.from_user.username
    user_display_name = f"@{username}" if username else f"{message.from_user.first_name} {message.from_user.last_name}".strip()

    # Отправка сообщения администраторам
    for admin_id in ADMIN_IDS:
        await bot.send_message(admin_id,
                               f"Пользователь {user_display_name} сообщил о проблеме:\n"
                               f"Маркетплейс: {marketplace}\n"
                               f"Номер телефона: {phone}\n"
                               f"Сообщение: {user_message}\n")
        if is_media and media_id:
            if message.photo:
                await bot.send_photo(admin_id, photo=media_id)
            elif message.video:
                await bot.send_video(admin_id, video=media_id)

    # Отправка подтверждения пользователю
    await message.answer("Ваша информация успешно отправлена. Спасибо за обратную связь!")

    # Сохранение данных в Excel файл
    try:
        wb = openpyxl.load_workbook('problems.xlsx')  # Попытка загрузить существующий файл
    except FileNotFoundError:
        wb = Workbook()  # Если файл не найден, создаем новый
        ws = wb.active
        ws.append(['Пользователь', 'Маркетплейс', 'Номер телефона', 'Сообщение'])
    else:
        ws = wb.active

    row_data = [user_display_name, marketplace, phone, user_message]
    ws.append(row_data)

    wb.save('problems.xlsx')

    # Отображаем главное меню
    await send_welcome_and_show_main_menu(message, state)



if __name__ == '__main__':
    executor.start_polling(dp, skip_updates=True)